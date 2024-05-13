

"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
B04_Python Statements
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

"""***********************************************************************
1. Continuation:
    using \
    please note that nothing should be after \

***********************************************************************"""

x=1+2+3\
+5+6+7\
+8+9+10
 
x #-->51


"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
B05 Python Logic
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

"""***********************************************************************
and是取大者的意思
**********************************************************************"""
3 and 5         #-->5
3 and 5 and 7   #--> 7

"""***********************************************************************
连续的大小判断
**********************************************************************"""
4< (3 and 5) <7 #--> True
4< (3 and 5) >4 #--> True
4< (3 and 5) >7 #--> False
# -*- coding: utf-8 -*-
"""
Created on Thu Oct 18 08:37:46 2018

@author: Danish
"""

import functools 
import time

def timer(func): 
    """Print the runtime of the decorated function""" 
    @functools.wraps(func) 
    def wrapper_timer(*args, **kwargs): 
        start_time = time.perf_counter() # 1 
        value = func(*args, **kwargs) # run your function 
        end_time = time.perf_counter() # 2 
        run_time = (end_time - start_time)*(10**6) # 3 
        print(f"Finished {func.__name__!r} in {run_time:.4f} m secs") 
        return value
    return wrapper_timer

report_bin={}
def reporter(func):
    """print the funct name of the decorated function"""
    @functools.wraps(func)
    def wrapper_reporter(*args,**kwargs):
        value=func(*args,**kwargs)
        return value
    report_bin[func.__doc__]=wrapper_reporter()
    return wrapper_reporter

@timer
@reporter
def print_numbers():
    "I am printing numbers"
    return 1*2




print_numbers()

"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
Combine lines with the same feature
example: combine risk node lines with the same AE ID
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

import pandas as pd

"""********************************************************************
1.Simple Case
********************************************************************"""
data=[['A','apple'],['A','avocado'],['B','banana'],['B','blueberry']]
df=pd.DataFrame(data,columns=['Initial','Fruit'])
"""
  Initial      Fruit
0       A      apple
1       A    avocado
2       B     banana
3       B  blueberry"""

output=df[['Initial']].drop_duplicates()
df=df.set_index('Initial')
def stack_lines(idx,df,col):
    return ('\n').join(list(df.loc[idx,col].values))
output['Fruit']=output['Initial'].apply(lambda x: stack_lines(x,df,'Fruit'))

"""********************************************************************
2. Multiple columns to combine
********************************************************************"""
data=[['A','apple','red'],
      ['A','avocado','green'],
      ['B','banana','yellow'],
      ['B','blueberry','blue']]
df=pd.DataFrame(data,columns=['Initial','Fruit','Color'])
"""
  Initial      Fruit   Color
0       A      apple     red
1       A    avocado   green
2       B     banana  yellow
3       B  blueberry    blue"""

output=df[['Initial']].drop_duplicates()
df=df.set_index('Initial')
def stack_lines(idx,df,col):
    return ('\n').join(list(df.loc[idx,col].values))

for col in {'Fruit','Color'}:
    output[col]=output['Initial'].apply(lambda x: stack_lines(x,df,col))
"""********************************************************************
3. Multiple columns to combine - different types
********************************************************************"""
data=[['A','apple','red',1],
      ['A','avocado','green',2],
      ['B','banana','yellow',3],
      ['B','blueberry','blue',4]]
df=pd.DataFrame(data,columns=['Initial','Fruit','Color','Qty'])
"""
  Initial      Fruit   Color  Qty
0       A      apple     red    1
1       A    avocado   green    2
2       B     banana  yellow    3
3       B  blueberry    blue    4"""


output=df[['Initial']].drop_duplicates()
df=df.set_index('Initial')
def stack_lines(idx,df,col):
    return ('\n').join(list(df.loc[idx,col].values.astype(str)))

for col in {'Fruit','Color','Qty'}:
    output[col]=output['Initial'].apply(lambda x: stack_lines(x,df,col))
    
"""********************************************************************
3. Final Generalized method
********************************************************************"""
data=[['A','apple','red',1],
      ['A','avocado','green',2],
      ['B','banana','yellow',3],
      ['B','blueberry','blue',4]]
df=pd.DataFrame(data,columns=['Initial','Fruit','Color','Qty'])
"""
  Initial      Fruit   Color  Qty
0       A      apple     red    1
1       A    avocado   green    2
2       B     banana  yellow    3
3       B  blueberry    blue    4"""


def combine_lines(df,key_col,combine_col_list):
    output=df[[key_col]].drop_duplicates()
    df=df.set_index(key_col)
    def stack_lines(idx,df,col):
        return ('\n').join(list(df.loc[idx,col].values.astype(str)))
    for col in combine_col_list:
        output[col]=output[key_col].apply(lambda x: stack_lines(x,df,col))
    return output

result=combine_lines(df,'Initial',['Fruit','Color','Qty'])

"""********************************************************************
4. Case study
********************************************************************"""
import os
myPath=r'C:\Users\Danish\Desktop\python sample'
os.chdir(myPath)
df=pd.read_excel('python sample.xlsx','Database') 

result=combine_lines(df,'Month',['Product','State','Sales $'])
result.to_excel('temp.xlsx')

"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
Find formulas in Speadsheets
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

import pandas as pd
import openpyxl
import os

myPath=r'C:\Users\Danish\Desktop\python sample'
os.chdir(myPath)

"""********************************************************************
Excel object basics
********************************************************************"""
#1 workbook object
wb_name='python sample.xlsx'
wb= openpyxl.load_workbook(wb_name) # get wb obj
wb.sheetnames                       # get all the sheet names

#2 worksheet object
ws=wb['formula']                    #get ws obj
active_ws=wb.active                 #get active ws
cl.row                              #row index such as 1
cl.column                           #col name such as 'A'

#3 cell object
cl=ws['A1']                         #type: openpyxl.cell.cell.Cell
cl=ws.cell(row=1,column=1)          #type: openpyxl.cell.cell.Cell
cl_value=ws['A1'].value             #type: str

"""********************************************************************
Loop through a range
********************************************************************"""
wb_name='python sample.xlsx'
wb= openpyxl.load_workbook(wb_name) # set workbook obj
ws=wb['formula']                    #set worksheet obj

formula_dict={}

for therow in range(1,4,1):
    for thecol in range(1,3,1):
        mycell=ws.cell(row=therow,column=thecol)
        if any (key_word in str(mycell.value) for key_word in ['+','-','=']):
            formula_dict[mycell]="'"+mycell.value

"""********************************************************************
Loop through entire sheet
********************************************************************"""
def formula_in_ws(wb_name,ws_name,key_word_list):
    wb= openpyxl.load_workbook(wb_name)
    ws_df=pd.read_excel(wb_name,ws_name,header=None)
    formula_dict={}
    
    ws_df=pd.read_excel(wb_name,ws_name,header=None)
    ws=wb[ws_name]
    row_range=ws_df.shape[0]
    col_range=ws_df.shape[1]
    
    for therow in range(1,row_range+1,1):
        for thecol in range(1,col_range+1,1):
            mycell=ws.cell(row=therow,column=thecol)
            if any (key_word in str(mycell.value) for key_word in key_word_list):
                formula_dict[mycell]="'"+mycell.value
    formula_df=pd.DataFrame.from_dict(formula_dict,orient='index',columns=['Formula'])
    return formula_df

results=formula_in_ws('python sample.xlsx','formula',['+','-','='])

"""********************************************************************
Loop through entire workbook
********************************************************************"""
def formula_in_wb(wb_name,key_word_list):
    wb= openpyxl.load_workbook(wb_name)
    ws_name_list=list(wb.sheetnames)
    formula_dict={}
    
    for ws_name in ws_name_list:
        ws_df=pd.read_excel(wb_name,ws_name,header=None)
        ws=wb[ws_name]
        row_range=ws_df.shape[0]
        col_range=ws_df.shape[1]
        for therow in range(1,row_range+1,1):
            for thecol in range(1,col_range+1,1):
                mycell=ws.cell(row=therow,column=thecol)
                if any (key_word in str(mycell.value) for key_word in key_word_list):
                    formula_dict[mycell]="'"+str(mycell.value)
    formula_df=pd.DataFrame.from_dict(formula_dict,orient='index')
    return formula_df

results=formula_in_wb('python sample.xlsx',['+','-','='])
results.to_excel('temp.xlsx')

"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
C03 Compare two DF for diff
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""
import pandas as pd
import os
myPath=r'C:\Users\Danish\Desktop\python sample'
os.chdir(myPath)


"""********************************************************************
Sample Data
Differences:
    1.column names: a1 vs a2
    2.index names: z1 vs z2
    3. values: 5 vs 10
********************************************************************"""
#1 df_a
data=[[1,2,3],[4,5,6],[7,8,9]]
df_a=pd.DataFrame(data,columns=['a1','b','c'],index=['x','y','z1'])
"""
   a1  b  c
x  1  2  3
y  4  5  6
z1  7  8  9"""
#2 df_b
data=[[1,2,3],[4,10,6],[7,8,9]]
df_b=pd.DataFrame(data,columns=['a2','b','c'],index=['x','y','z2'])
"""
    a2   b  c
x    1   2  3
y    4  10  6
z2   7   8  9"""

"""********************************************************************
Define my function:
    Based on the 'option' selected by the user, this function can return
    columns only in one DF, indexes only in one DF, or different values
    in the same column and index
********************************************************************"""
#1 find columns only in df_a
def df_comparison(df_a,df_b,option):
    
    col_diff=list(set(df_a.columns.values)-set(df_b.columns.values))
    idx_diff=list(set(df_a.index.values)-set(df_b.index.values))
    col_same=list(set(df_a.columns.values)& set(df_b.columns.values))
    idx_same=list(set(df_a.index.values)& set(df_b.index.values))
    
    if option=='col':
        return df_a.loc[:,col_diff]
    elif option=='idx':
        return df_a.loc[idx_diff,:]
    elif option== 'df':
        df_a_com=df_a.loc[idx_same,col_same]
        df_b_com=df_b.loc[idx_same,col_same]
        
        boolean_df=(df_a_com!=df_b_com)
        diff_df_a=((df_a_com[boolean_df]).dropna(how='all',axis=0)).dropna(how='all',axis=1)
        diff_df_b=((df_b_com[boolean_df]).dropna(how='all',axis=0)).dropna(how='all',axis=1)
        
        diff_df_a['Source']='a'
        diff_df_b['Source']='b'
        
        diff_df=pd.concat([diff_df_a,diff_df_b],axis=0)
        diff_df=diff_df.sort_index()
        
        return diff_df
    else:
        print("option can only be'col','idx' or 'df'.")

"""********************************************************************
Testing my function:
********************************************************************"""
#1 Find cols only in df_a
df_comparison(df_a,df_b,'col')
"""
   a1
x  1
y  4
z1  7"""

#2 Find cols only in df_b
df_comparison(df_b,df_a,'idx')

"""
    a2  b  c
z2   7  8  9"""

#3 Find different values in col/row that exist in both
df_comparison(df_a,df_b,'df')
"""
      b Source
y   5.0      a
y  10.0      b"""
"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
C04 Appearance Counter
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

"""********************************************************************
Count appearance by using DICT.setdefault

Use:
    automatically setting d[key] to the list only when it's unset.
********************************************************************"""

message = 'abbcccddd'

counter = {}
counter.setdefault('a', 0)
counter['a'] = counter['a'] + 1

for character in message:
    counter.setdefault(character, 0)
    counter[character] = counter[character] + 1


"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
D01 Class 101
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

"""********************************************************************

Define a class which has at least two methods:
getString: to get a string from console input
printString: to print the string in upper case.
Also please include simple test function to test the class methods.

********************************************************************"""

class InputOutString(object): #1 use "class" to start & with "object" in ()
    def __init__(self):       #2 always first define __init__(self)
        self.s = ""           #3 define the basic property with self.XXXX

    def getString(self):       #1 define a function with (self)
        self.s = input() 

    def printString(self):
        print (self.s.upper())

strObj = InputOutString()     #1 To use a class, just assign it to a var
strObj.getString()            #2 To use its function, var.function()
strObj.printString()


"""********************************************************************
理解class和其中method和attribute的用法 - 简单例子
********************************************************************"""
class maomi(object):
    def __init__(self):
        self.name='bad cat'
    def get_name(self):
        self.name=input()
    def get_age(self):
        self.age=int(input())
    def double_age(self):
        self.double=self.age*2
    def triple_age(self):
        return self.age*3
    
#1 首先把一个变量定义为这个class
maomao=maomi()

# Call attribute
maomao.age
"""AttributeError: 'maomi' object has no attribute 'age'
这个说明你没有用get_age求age这个attribute前，maomao是没有这个attribute的
"""

maomao.name
"""Out[193]: 'bad cat'
因为我在init里面定义了self.name='bad cat'所以，这个是天然就有的"""

# Call method to define attribute and then call attribute

maomao.get_name()
"""要我输入，我就输入了：
mao mao, 现在再看maomao.name"""

maomao.name
"""Out[195]: 'mao mao'"""

maomao.get_age() # I input 38
maomao.age       # --> 38

maomao.double
"""AttributeError: 'maomi' object has no attribute 'double'
这个和上面age出现的错误是一样的，你没有run double_age()前是没有double这个
attribute的"""

maomao.double_age()
maomao.double    #--> 76


"""********************************************************************
理解class和其中method和attribute的用法 - case study

Question:
A robot moves in a plane starting from the original point (0,0). 
The robot can move toward UP, DOWN, LEFT and RIGHT with a given steps. 
The trace of robot movement is shown as the following:

UP 5,DOWN 3,LEFT 3,RIGHT 2

The numbers after the direction are steps. 
Please write a program to compute the distance from current position 
after a sequence of movement and original point. 
If the distance is a float, then just print the nearest integer.

Example:
If the following tuples are given as input to the program:
UP 5
DOWN 3
LEFT 3
RIGHT 2

Then, the output of the program should be:
2


********************************************************************"""

class movement(object):
    def __init__(self):
        self.myMove={'UP':0,'DOWN':0,'LEFT':0,'RIGHT':0}
    
    def get_movement(self):
        myMove=self.myMove
        while True:
            s=input()
            if s:
                mySplit=s.split(" ")
                myDir=mySplit[0]
                myLength=int(mySplit[1])
                myMove[myDir]=myMove[myDir]+myLength
            else:
                break
        self.myMove=myMove    
        
    def get_location(self):
        myLoc={'X':0,'Y':0}
        myLoc['Y']=myLoc['Y']+self.myMove['UP']-self.myMove['DOWN']
        myLoc['X']=myLoc['X']+self.myMove['RIGHT']-self.myMove['LEFT']
        self.loc=myLoc
        
    def get_distance(self):
        myLoc=self.loc
        raw_dist=((myLoc['X'])**2+(myLoc['Y'])**2)**(0.5)
        self.distance=int(round(raw_dist,0))
        
    
mytest=movement()

mytest.myMove #返回default{'UP': 0, 'DOWN': 0, 'LEFT': 0, 'RIGHT': 0}
mytest.get_movement() # run method to provide input
mytest.myMove #{'UP': 10, 'DOWN': 2, 'LEFT': 7, 'RIGHT': 1}

mytest.get_location()
mytest.loc #{'X': -6, 'Y': 8}

mytest.get_distance()
mytest.distance #-->10

"""我现在换另一个方法,想说明顺序对于class method的重要性"""
maotest=movement()
maotest.get_movement()
maotest.get_distance() #AttributeError: 'movement' object has no attribute 'loc'
#因为计算distance是建立在计算location之后的，所以不能在没有run get_location之前
# run get_distance


"""********************************************************************
如何定义需要外部输入的类

********************************************************************"""
class Circle(object):
    def __init__(self,r):
        self.radius=r
    def area(self):
        return (self.radius)**2*3.1415926

myCircle=Circle(4)
myCircle.radius
myCircle.area()




"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
Decorator
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

"""***********************************************************************
1. Simple Decorator
Put simply: decorators wrap a function, modifying its behavior.
在下面的例子就可以看到，我原本的function say_whee()只能返回"Whee!"，
但是my_decorator改变了它：前后还加上了2个print。这个很有启发意义
我有过很多类似的test，可能就是input不同，逻辑是一样的。用wrapper是不是说我就
可以不用重复那么多了呢？
***********************************************************************"""

#—————————————————————————————————————————————————————————————————————————
def my_decorator(func):
    def wrapper():
        print("Something is happening before the function is called.")
        func()
        print("Something is happening after the function is called.")
    return wrapper

def say_whee():
    print("Whee!")
say_whee = my_decorator(say_whee)

# If I run this code, say_whee is a function
say_whee 
"""
<function __main__.my_decorator.<locals>.wrapper()>"""

say_whee()
"""
Something is happening before the function is called.
Whee!
Something is happening after the function is called."""

#—————————————————————————————————————————————————————————————————————————
from datetime import datetime

def not_during_the_night(func):
    def wrapper():
        if 7 <= datetime.now().hour < 22:
            func()
        else:
            pass  # Hush, the neighbors are asleep
    return wrapper

def say_whee():
    print("Whee!")

say_whee = not_during_the_night(say_whee)
say_whee()

"""***********************************************************************
2. Use @decorator re-write的code above
语法就是
1.先定义一个function，比如叫my_decor(func) -需要def def return的形式
2. @my_decor
3. 定义这个func是什么，比如叫my_func()
4. call my_func()就是2个函数合并的结果
***********************************************************************"""
# ______________________Re-write the first one ___________________________

def my_decorator(func):
    def wrapper():
        print("Something is happening before the function is called.")
        func()
        print("Something is happening after the function is called.")
    return wrapper

@my_decorator
def say_whee():
    print("Whee!")

"""我现在run say_whee() 直接得到的就是

Something is happening before the function is called.
Whee!
Something is happening after the function is called.

前面的code里需要写 say_whee = my_decorator(say_whee)"""

# ______________________Practice myself ___________________________

def reporting(func):
    def wrapper():
        print('Your test result is {}'.format(func()))
    return wrapper

@reporting
def test():
    myInput=int(input())
    if myInput%2==0:
        return 'Even'
    else:
        return "Odd"

test()
"""上面有几个注意的点：func后面必须要有()，否则返回的结果不是function的结果
而是function本人，比如<function test at 0x0000029926CBB8C8>"""

"""***********************************************************************
3. Decorator that allows arguments in func
***********************************************************************"""
#——————————————————— 没有参数的情况————————————————————————————————————————
def do_twice(func):
    def wrap():
        func()
        func()
    return wrap

@do_twice
def print_name():
    print ("Qiusi")

print_name()
"""
Qiusi
Qiusi"""

#——————————————————— 加上参数的情况(行不通的方法)———————————————————————————

def do_twice(func):
    def wrap():
        func()
        func()
    return wrap

@do_twice
def print_name(num):
    print (num)

print_name('Qiusi')
"""
TypeError: wrap() takes 0 positional arguments but 1 was given
我们会得到一个error message，因为decorator是不接受arguments的，
如果我们需要arg，请看下面的方法
"""
#——————————————————— 加上参数的情况(可行的方法)———————————————————————————
"""The solution is to use *argsand **kwargsin the inner wrapper function. 
Then it will accept an arbitrary number of positional and keyword arguments. """

def do_twice(func):
    def wrap(*args,**kwargs):
        func(*args,**kwargs)
        func(*args,**kwargs)
    return wrap

@do_twice
def print_name(num):
    print (num)

print_name('Qiusi')

"""***********************************************************************
4. Decor Returns Value
***********************************************************************"""
#——————————————————— Return value的情况(行不通的方法)———————————————————————
def do_twice(func):
    def wrap(*args,**kwargs):
        func(*args,**kwargs)
        func(*args,**kwargs)
    return wrap

@do_twice
def return_greeting(name): 
    print("Creating greeting") 
    return f"Hi {name}"

return_greeting('Qiusi')

"""
我期待的结果是这样的：
Creating greeting
Hi Qiusi
Creating greeting
Hi Qiusi

但是结果是这样的，return后面的内容没有反应出来
Creating greeting
Creating greeting

所以我们还是要修正一下decorator"""

#——————————————————— Return value的情况(可行的方法)—————————————————————————
def do_twice(func): 
    def wrapper_do_twice(*args, **kwargs):
        func(*args, **kwargs)
        func(*args, **kwargs)
        return func(*args, **kwargs) #改变是要加上return这行
    return wrapper_do_twice

@do_twice
def return_greeting(name): 
    print("Creating greeting") 
    return f"Hi {name}"

return_greeting('Qiusi')

"""结果是这样的：
Creating greeting
Creating greeting
Creating greeting
Out[95]: 'Hi Qiusi'

其实还是和我想的不太一样，是这样分配input和output的
func(*args, **kwargs)           --> Creating greeting
func(*args, **kwargs)           --> Creating greeting
return func(*args, **kwargs) 
                                --> Creating greeting
                                --> Out[95]: 'Hi Qiusi'
所以就是说return会既返回action又返回value
"""



""""**********************************************************************
Introspection
***********************************************************************"""
#1 function性质被wrapper取代了
"""如果我们直接用python的__name__或者help什么的在被wrap的function上，
我们得到的是decorator的性质，而不是function的性质"""

return_greeting.__name__
"Out[]: 'wrapper_do_twice'"

help(return_greeting)
"""
Help on function wrapper_do_twice in module __main__:
wrapper_do_twice(*args, **kwargs)"""

#2 如果我们想得到function本身的性质，我们需要在decor做手脚
 
import functools
def do_twice(func): 
    @functools.wraps(func) #加入functools.wraps(func)
    def wrapper_do_twice(*args, **kwargs):
        func(*args, **kwargs)
        func(*args, **kwargs)
    return wrapper_do_twice

@do_twice
def return_greeting(name): 
    print("Creating greeting") 

return_greeting('Qiusi')


return_greeting.__name__
"Out[]: 'return_greeting'"

help(return_greeting)
"""
Help on function return_greeting in module __main__:
return_greeting(name)"""


"""***********************************************************************
Syntax template for decorator
***********************************************************************"""

import functools
def decorator(func): 
    @functools.wraps(func) 
    def wrapper_decorator(*args, **kwargs): 
        # Do something before 
        value = func(*args, **kwargs) 
        # Do something after 
        return value 
    return wrapper_decorator

"""***********************************************************************
Example 1 - Time Functions
***********************************************************************"""
import functools 
import time

def timer(func): 
    """Print the runtime of the decorated function""" 
    @functools.wraps(func) 
    def wrapper_timer(*args, **kwargs): 
        start_time = time.perf_counter() # 1 
        value = func(*args, **kwargs) # run your function 
        end_time = time.perf_counter() # 2 
        run_time = end_time - start_time # 3 
        print(f"Finished {func.__name__!r} in {run_time:.4f} secs") 
        return value 
    return wrapper_timer

@timer 
def waste_some_time(num_times): 
    for i in range(num_times): 
        sum([i**2 for i in range(10000)])

waste_some_time(100)

time.perf_counter.__doc__

"""***********************************************************************
Example 2 - Debugging code
***********************************************************************"""
#_____________________Reusable debug code_________________________________
import functools
def debug(func): 
    """Print the function signature and return value""" 
    @functools.wraps(func)
    def wrapper_debug(*args, **kwargs):
        args_repr = [repr(a) for a in args] # 1 
        kwargs_repr = [f"{k}={v!r}" for k, v in kwargs.items()] # 2
        signature = ", ".join(args_repr + kwargs_repr) # 3 
        print(f"Calling {func.__name__}({signature})") 
        value = func(*args, **kwargs) 
        print(f"{func.__name__!r} returned {value!r}") # 4 
        return value 
    return wrapper_debug

#____________________Example one: return value____________________________
@debug
def pricing(footage,age,school_score):
    est_price=footage*300-age*1000+school_score*20000
    return est_price

pricing(1167,10,8)
"""
Calling pricing(1167, 10, 8)
'pricing' returned 500100
Out[7]: 500100"""

#____________________Example two: action only_____________________________
@debug
def personal(name,age, address):
    print (f'My name is {name}. I am {age} years old. I live in {address}.')

personal('doudian',1,'Mamaroneck')
"""
Calling personal('doudian', 1, 'Mamaroneck')
My name is doudian. I am 1 years old. I live in Mamaroneck.
'personal' returned None"""

#___________________Example 3: apply debug to other code__________________
# factorial 阶乘的意思
import math 
# Apply a decorator to a standard library function 
math.factorial = debug(math.factorial)
def approximate_e(terms=18): 
    return sum(1 / math.factorial(n) for n in range(terms))

approximate_e(6)
"有个问题就是，为啥我要把这个debug加进来？就是起到展示每步结果的作用"

"""***********************************************************************
Slowing down code - 延时器，倒数器
***********************************************************************"""
import functools 
import time

def slow_down(func):
    """Sleep 1 second before calling the function""" 
    @functools.wraps(func)
    def wrapper_slow_down(*args, **kwargs): 
        time.sleep(1) 
        return func(*args, **kwargs) 
    return wrapper_slow_down

@slow_down 
def countdown(from_number):
    if from_number < 1: 
        print("Liftoff!") 
    else: 
        print(from_number) 
        countdown(from_number - 1)

countdown(20)

"""**********************************************************************
Registering Plugins
这个对我的启发很大，比如我要记录发现的Exception,可以给我的每个testing function
上register decor，然后能记录所有exception在一个dictionary

***********************************************************************"""
import random 

#_____________step 1 create a DICT and decor______________________________
PLUGINS = dict()
def register(func): 
    """Register a function as a plug-in""" 
    PLUGINS[func.__name__] = func 
    return func

#____________step2 apply decor to a function______________________________
@register 
def say_hello(name): 
    return f"Hello {name}"

# let me test it
say_hello('Qiusi')  # --> 'Hello Qiusi'
PLUGINS         #--> ]: {'say_hello': <function __main__.say_hello(name)>}

@register 
def be_awesome(name): 
    return f"Yo {name}, together we are the awesomest!"

be_awesome('Dudu') #-->'Yo Dudu, together we are the awesomest!'
PLUGINS
"""{'say_hello': <function __main__.say_hello(name)>,
 'be_awesome': <function __main__.be_awesome(name)>}"""

#______________step 3 randomly pick up a function and run_________________

list(PLUGINS.items()) # convert DICT to a LIST
random.choice(list(PLUGINS.items())) # pick up one random from the list

def randomly_greet(name): 
    greeter, greeter_func = random.choice(list(PLUGINS.items())) 
    print(f"Using {greeter!r}") # print out the function name
    return greeter_func(name)   # return function results

randomly_greet('Maomao')

"""***********************************************************************
User Login

这个flask好像是set up website用的，我现在还是放着
***********************************************************************"""
from flask import Flask, g, request, redirect, url_for 
import functools 
app = Flask(__name__)

def login_required(func): 
    """Make sure user is logged in before proceeding""" 
    @functools.wraps(func) 
    def wrapper_login_required(*args, **kwargs):
        if g.user is None: 
            return redirect(url_for("login", next=request.url)) 
        return func(*args, **kwargs) 
    return wrapper_login_required
@app.route("/secret") 
@login_required 

def secret():
"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
D03 Decorator 102
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

"""***********************************************************************
Some commonly used decorators that are even built-ins in Python are :
    @classmethod, 
    @staticmethod, and 
    @property. 
    
The @classmethodand @staticmethod decorators are used to define methods 
inside a class namespace that are not connected to a particular instance 
of that class. 

The @propertydecorator is used to customize getters and settersfor class 
attributes. 

***********************************************************************"""

"""***********************************************************************
Decorating Classes
***********************************************************************"""

#_______________________import the class and module_______________________
import sys
import os
os.chdir(r'C:\Users\Danish\Desktop\python sample')
sys.path.append(os.getcwd())
from Decorator import timer
from Decorator import debug

"""***********************************************************************
Decorate a class
***********************************************************************"""

class TimeWaster: 
    @debug 
    def __init__(self, max_num): 
        self.max_num = max_num
    @timer 
    def waste_time(self, num_times): 
        for _ in range(num_times):
            sum([i**2 for i in range(self.max_num)])

tw = TimeWaster(1000) 
"""
Calling __init__(<time_waster.TimeWaster object at 0x7efccce03908>, 1000) 
'__init__' returned None"""


tw.waste_time(999) 
"Finished 'waste_time' in 0.3376 secs"

"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
D04 Function
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""


"""********************************************************************
1. Print documentation for built-in functions
********************************************************************""" 
print (abs.__doc__)
print (int.__doc__)
print (input.__doc__)

"""********************************************************************
2. Create documentation for your own functions

syntax:
    create: 紧贴着function name下面用单引号或三引号书写documentation
    ***注意：如果不是紧贴，就不会被当作documentation，见下面的例子3
    
    call:functionname.__doc__.这个也可以用于查看别人或者built-in function
********************************************************************""" 
#____________________1. Multiple line documentation_______________________
def square(num):
    '''Return the square value of the input number.
    The input number must be integer.
    '''
    return num ** 2
print (square.__doc__)
"""
[out]
Return the square value of the input number.
    The input number must be integer."""

#_________________2. One line doc can use single quotation________________
def maomi_test(catname):
    " All the cats are good cats."
    print (catname + " is a good cat.")

maomi_test('maomao')
print(maomi_test.__doc__)
"""
[out]
maomao is a good cat.
 All the cats are good cats."""

#_________________3. What if I did not quote right below?________________
def maomi_test(catname):
    print (catname + " is a good cat.")
    " All the cats are good cats."

maomi_test('maomao')
print(maomi_test.__doc__)

"""[out]
maomao is a good cat.
None"""

# 可见不紧贴就这样就不能返回你的引号的内容

"""********************************************************************
3. Self-circulated functions
这个是我最近知道的，我其实可以在自己创造的function中，call这个function本人
从而达到一个循环的效果。
********************************************************************""" 
#——————————这是一个求阶乘的例子—————————————————————————————————————————————
def circle(n):
    if n>0:
        result=n*circle(n-1) # cycled
    elif n==0:
        result=1
    else:
        result= "n cannot be a negative number" 
    return result

circle(-1)



"""********************************************************************
1. How to take multiple-line input (without ending the input)
如何接受多行的输入？

Question 9
Level 2

Question:
Write a program that accepts sequence of lines as input and prints the 
lines after making all characters in the sentence capitalized.

Suppose the following input is supplied to the program:
Hello world
Practice makes perfect

Then, the output should be:
HELLO WORLD
PRACTICE MAKES PERFECT
********************************************************************"""
def collect_input():
    lines = []
    while True:
        s = input()
        if s:
            lines.append(s)
        else:
            break
    return lines

inPut=collect_input()

"""***********************************************************************
1. Raise runtime error
我也不知道这个对我能有什么用，先记下来
应该是编写function的时候可以提示用户问题所在
应该研究一下raise的用法
***********************************************************************"""
raise RuntimeError('Maomao is the best')

RuntimeError.__doc__

#————————————————————————KeyError is the same—————————————————————————————

def learn_key_error():
    myInput=input()
    if myInput in ['mao','maomi','maomao']:
        print ('My name is '+myInput)
    else:
        raise KeyError('Mao rejects your input.')

learn_key_error()    

"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
L03 Sub Class
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

"""********************************************************************
1.Sub Class basic - inheritance
********************************************************************"""
#1 Define a class
class Shape(object):
    def __init__(self):
        pass
    def document(self):
        print ('Shape is shape.')

myShape=Shape()
myShape.document()  #--->Shape is shape.

#2 Define a sub-class using the class defined above

class Triangle(Shape):
    pass

subShape=Triangle()
subShape.document() #--->Shape is shape.
# even though I put nothing in the Triangle, it inherits function from Shape

"""********************************************************************
2.Sub Class basic - modify what it inherits
********************************************************************"""
class Animal(object):
    def __init__(self):
        pass
    def feature1(self):
        print ('Animals are lovely.')
    def feature2(self):
        print ('Animals are cute.')

class Cat(Animal):
    def feature2(self):
        print ('Cats are cute.') # I only modify feature2 function

myCat=Cat()
myCat.feature1() #--> Animals are lovely.
myCat.feature2() #--> Cats are cute.

"""********************************************************************
3.Check object name
********************************************************************"""
# use the example above

myCat           #--> <__main__.Cat at 0x29927091630>
myAnimal=Animal()
myAnimal        #--> <__main__.Animal at 0x2992709a0b8>



"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
User-defined Module:
    1. How to create a user-defined module
    2. How to import user-defined module for use

For this exercise, I created a file named "Decorator"
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

"""***********************************************************************
Create a user-defined moduel:
    1. Below is copied pasted from the "Decorator" for illustration
    2. we need to define a class or a function first
    3. if you develop a class, we need to define functions under the class
    4. if you develop a function, then no else to be done
    5. save the module file in the cwd
如果你的function是在class下面，你就要从file import class，然后在class.func
如果你的function就是独立的，你就直接从file import func
***********************************************************************"""
"""
import functools 
import time

def timer(func): 
    'Print the runtime of the decorated function' 
    @functools.wraps(func) 
    def wrapper_timer(*args, **kwargs): 
        start_time = time.perf_counter() # 1 
        value = func(*args, **kwargs) # run your function 
        end_time = time.perf_counter() # 2 
        run_time = end_time - start_time # 3 
        print(f"Finished {func.__name__!r} in {run_time:.4f} secs") 
        return value 
    return wrapper_timer

class DecoratorClass(object):
    def myTest():
        print ('Congratulations!')
""" 
"""***********************************************************************
Import a user-defined module for use
1. Make sure the current folder is the folder where the module is
2. use sys.path.append() to add the current folder to sys.path
3. from thefilename import theclassname
***********************************************************************"""
import sys
import os
os.chdir(r'C:\Users\Danish\Desktop\python books')
sys.path.append(os.getcwd())

#_____________Import a class from a file and use its function_____________
from DecoratorFile import DecoratorClass as dct

dct.myTest() #-> Congratulations!

#___________________Import a function from a file__________________________
from DecoratorFile import timer

@timer
def drill():
    print ('meow')

drill()

"""
meow
Finished 'drill' in 0.0001 secs"""

"""***********************************************************************
namedtuples to define a class
***********************************************************************"""

from collections import namedtuple       # step 1 import namedtuple
Car = namedtuple('Car', 'color mileage') # step 2 define class and attributes
# Car is class name, color and milage are attributes
my_car = Car('red', 3812.4)              # step 3 grant values  

my_car                                   #--> Car(color='red', mileage=3812.4)              
my_car.color                             #--> 'red'
my_car.mileage                           #--> 3812.4
my_car.__doc__                           #--> 'Car(color, mileage)'
my_car.__class__                         #--> __main__.Car


#____________Try another example__________________________________________
Cat=namedtuple('Kitty',['age','weight'])
my_cat=Cat(3,5)
my_cat.age                               #--> 3
my_cat.weight                            #--> 5
my_cat.__class__                         #-->  __main__.Kitty
my_cat                                   #--> Kitty(age=3, weight=5)

#_________________This is tuple, immutable________________________________
my_cat.age=6
"AttributeError: can't set attribute"
"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
L01 Dictionary
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

"""********************************************************************
Creat a DICT
********************************************************************"""
#1 Create DICT using {key:value}
{'a':1,'b':2,'c':3}
"[out]-->{'a': 1, 'b': 2, 'c': 3}"

#2 Crate DICT using dict(key=value)
dict(a=1,b=2,c=3) 
"[out]-->{'a': 1, 'b': 2, 'c': 3}"

#3 Create DICT using zip
myZip=zip(['a','b','c'],range(1,4))   # this step creates a zip object
dict(myZip)
"[out]-->{'a': 1, 'b': 2, 'c': 3}"

#4 使用zip，如果你的key和value的数量不一样，python就取短者
myZip=zip(['a','b','c'],range(1,10))
dict(myZip)
"[out]-->{'a': 1, 'b': 2, 'c': 3}"

myZip=zip(['a','b','c','d','e'],range(1,4))
dict(myZip)
"[out]-->{'a': 1, 'b': 2, 'c': 3}"

# Create an empty dict
myEmpty=dict()
type(myEmpty) # --> Out[]: dict
"""********************************************************************
Retreat value/key from DICT
********************************************************************"""
#2 Get All DICT key, value or both

myDict.keys()               #dict_keys(['a', 'b', 'c'])
myDict.values()             #dict_values([1, 2, 3])
list(myDict.values())       #[1, 2, 3]
list(myDict.items())        #[('a', 1), ('b', 2), ('c', 3)]

#3 Get Certain Dict value
myDict['a']                 #Out[]: 1
myDict.get('a')             #Out[]: 1
myDict.get('a','Not Found') #Out[]: 1
myDict.get('e','Not Found') #Out[]: 'Not Found'

"""********************************************************************
Add new items to a dict
********************************************************************"""
myDict={'a':1,'b':2,'c':3}
myDict.update({'d':4})
myDict #{'a': 1, 'b': 2, 'c': 3, 'd': 4}

"""********************************************************************
Add new items
Use DICT.setdefault
********************************************************************"""
# 如果发现新key就自动加入dictionary，如果已经set了，就忽略它
message = 'abbcccddd'

counter = {}
counter.setdefault('a', 0)
counter['a'] = counter['a'] + 1

for character in message:
    counter.setdefault(character, 0)
    counter[character] = counter[character] + 1
"""********************************************************************
DICT's statistics
********************************************************************"""
myDict=dict(a=1,b=2,c=3,d=4)
sum(myDict.values()) #--> 10

min(myDict.values()) #--> 1

import statistics as stat
stat.stdev(myDict.values()) #--> 1.290994

"""********************************************************************
SORT DICT into list
********************************************************************"""
myDict=dict(z=1,b=2,x=3,d=4)
sorted(myDict.items())          #1 sort the entire dictionary per keys
"""[('b', 2), ('d', 4), ('x', 3), ('z', 1)]"""

sorted(myDict.keys())           #2 sort keys only
"""['b', 'd', 'x', 'z']"""

sorted(myDict.values())         #3 sort values only
"""[1, 2, 3, 4]"""

from operator import itemgetter #4 sort entire dict per values, reverese
sorted(myDict.items(),key=itemgetter(1,0),reverse=True)


sorted(myDict.items(), key=lambda x: x[1])
"[('z', 1), ('b', 2), ('x', 3), ('d', 4)]"
                                #5 sort entire dict per values, 2nd way
                                
sorted(myDict.items(), key=lambda x: -x[1])
"[('d', 4), ('x', 3), ('b', 2), ('z', 1)]"
                                #6 sort entire dict per values, reverese
"""********************************************************************
Merge two dictionaries
********************************************************************"""
#_______________dict.update(dict)的方法会破坏原有的dict____________________
dict1={'a':1,'b':2}
dict2={'b':3,'c':3}
dict3=dict1.update(dict2) #这样不行，因为dict3是个method，dict1还是被改变了

#_———————————————————————正确的方法是用{**dict1,**dict2}————————————————————
#右边的override左边的
{**dict1,**dict2} #--> {'a': 1, 'b': 3, 'c': 3}
{**dict2,**dict1} #--> {'b': 2, 'c': 3, 'a': 1}

#——————————————————Multiple dicts will be the same————————————————————————
dict1={'a':1,'b':2}
dict2={'b':3,'c':3}
dict3={'b':4,'c':5}

{**dict1,**dict2,**dict3} #—> {'a': 1, 'b': 4, 'c': 5}
#顺序是dict2 override dict1然后 dict3再override

"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
L02 List
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

"""********************************************************************
1.Create list
syntax:
    [value,value]
********************************************************************"""
#1 List use "[]" to contain and "," to separate
# Its components can be strings, numbers, other lists, tuples, dict..
['cat', 'bat', 'rat', 'elephant']

"""********************************************************************
2.Append/Insert New element
syntax:
    LIST.append(value)
    LIST.insert(location,value)
********************************************************************"""

#1. Append at the end
myList=['cat', 'bat', 'rat', 'elephant']
myList.append('moose')
"""
myList
Out[31]: ['cat', 'bat', 'rat', 'elephant','moose']"""

####################### WARNING! #######################
"""myList.append('moose') is a method, not a list.
So if you assign it to a variable, for example
newList=myList.append('moose')
you will see newList is a None Type and has no value
and myList is appended with 'moose'.

What does this suggest?
It tells us that Using "append" or "insert" (see below)
will modify the original list, whether or not you assign
it to a new variable. 

"""
########################################################

#2. Insert at certain location
myList=['cat', 'bat', 'rat', 'elephant']
myList.insert(1, 'chicken')
"""
myList
Out[38]: ['cat', 'chicken', 'bat', 'rat', 'elephant']
"""

"""********************************************************************
3. Select/Slice a list
syntax:
    LIST[location]
    LIST[start(inclusive):end(exclusive):step size]
********************************************************************"""
#1 select one element
['cat', 'bat', 'rat', 'elephant'][3] # output ->'elephant'

#2 select with start, end and step LIST[start:end:step]
# NOTE that, in Python, left inclusive right exclusive

# All parameters are positive
[1,2,3,4,5,6,7][3::]               #output-> [4, 5, 6, 7]
[1,2,3,4,5,6,7][1:3]               #output-> [2, 3]
[1,2,3,4,5,6,7][1:6:2]             #output-> [2, 4, 6]
[1,2,3,4,5,6,7,8,9][1::2]          #output->[2, 4, 6, 8]

# Negative parameters
[1,2,3,4,5,6,7][-2::]              #Start from -2 position [6, 7]
[1,2,3,4,5,6,7][::-1]              #Reversed order  [7, 6, 5, 4, 3, 2, 1]
[1,2,3,4,5,6,7][::-2]              #Reserved ->[7, 5, 3, 1]
[1,2,3,4,5,6,7,8,9][-2::-2]        #Start from -2 and reverse [8, 6, 4, 2]

"""********************************************************************
4. Delete elements from a list
syntax:
    del LIST[location]
    LIST.remove(value)
********************************************************************"""
#1 Delete per location
myList=['cat', 'bat', 'rat', 'elephant']
del myList[2]
"""
myList
Out[44]: ['cat', 'bat', 'elephant']"""

#2 Delete per value
myList=['cat', 'bat', 'rat', 'elephant']
myList.remove('cat')
"""
myList
Out[46]: ['bat', 'rat', 'elephant']"""

"""********************************************************************
5.Combine/Multiple lists
syntax:
    LIST + LIST
    LIST*#
********************************************************************"""
#1 use "+" to combine two list
[1, 2, 3] + ['A', 'B', 'C']
"""
[1, 2, 3, 'A', 'B', 'C']"""

#2 use "*3" to repeat
['X', 'Y', 'Z'] * 3
"""
['X', 'Y', 'Z', 'X', 'Y', 'Z', 'X', 'Y', 'Z']"""


"""********************************************************************
6.Assign values using List
syntax:
    variable1,variable2,variable3=[value1,value2,value3]
********************************************************************"""
size, color, disposition = ['fat', 'black', 'loud']
"""
In [2]: size
Out[2]: 'fat'

color
Out[3]: 'black'

disposition
Out[4]: 'loud'"""

"""********************************************************************
6.Flatten nested List
flat_list=[item for sublist in nested_list for item in sublist]
********************************************************************"""
nested=[[1,2,3],[4,5,6]]
[element for nest in nested for element in nest]

"""
[1, 2, 3, 4, 5, 6]"""

"""********************************************************************
7.Sort a List
syntax:
    LIST.sort()
    LIST.sort(reverse=True)
    sorted(LIST)
********************************************************************"""
#1 LIST.sort() -> permantly change
[1,3,2,4].sort()   # --> Nothing, because .sort() is a method

myList=[1,3,2,4]
myList.sort()
myList             # --> [1,2,3,4]

#1.1 Descending
myList.sort(reverse=True)
myList             # --> [4,3,2,1]

#2 sorted(LIST) --> display in a sorted way, but not change the original
myList=[1,3,2,4]
sorted(myList)     # --> [1, 2, 3, 4]
myList             # --> [1,3,2,4]

"""********************************************************************
8.Sort a List with key
syntax:
    LIST.sort(key=function name)

********************************************************************"""
#1 set simple keys
myList=['a', 'z', 'A', 'Z']
myList.sort()
myList              #-->['A', 'Z', 'a', 'z']

myList=['a', 'z', 'A', 'Z']
myList.sort(key=str.lower)
myList              #-->['a', 'A', 'z', 'Z']

#2 set keys with self-defined function

myList=[10,2,3,4]
def mySortStd(num):
      return (num**2)*(-1)*num
sorted(myList, key=mySortStd) # -->[10, 4, 3, 2]

#4 primary and secondary sorting critieria 

mylist=[('John', '20', '90'), ('Jony', '17', '91'), ('Jony', '17', '93'), 
('Json', '21', '85'), ('Tom', '19', '80')]

from operator import itemgetter
mylist.sort(key=itemgetter(2,0,1))

mylist
"""
[('Tom', '19', '80'),
 ('Json', '21', '85'),
 ('John', '20', '90'),
 ('Jony', '17', '91'),
 ('Jony', '17', '93')]"""

"""********************************************************************
9.List statistics
syntax:
    
********************************************************************"""
myList=[1,23,45,230]

sum(myList)             #--> 299
min(myList)             #--> 1

import statistics as stat
stat.stdev(myList)      #--> 105.0472116082415

"""********************************************************************
10. One-line list with if else
syntax:
    ["value_for_True" if "condition" else "value_for_False" for x in "loop" ]
    
********************************************************************"""
full_list=[1,2,3,4,5,6]
["Even" if x%2==0 else "Odd" for x in full_list ]
"""['Odd', 'Even', 'Odd', 'Even', 'Odd', 'Even']"""


"""********************************************************************
11. Bin a list into groups using pd.qcut
先解释一下这个东西是什么用，比如我想把一个班的学生分段，成绩最差25%一段，成绩25%
到75%之间一段，75%-100%一段。这时候可以用pandas的qcut功能

我觉得qcut是quantile cut的意思，不过任何percentage都可以

syntax:
    pd.qcut(ser,q=q_list,labels=label_list)
********************************************************************"""
#——————————————————Example 1: tell me your range —————————————————————————
import pandas as pd
#1 下面的list是我学生的分数，不需要sort
ser=[59,62,83,84,95,76,89,99]

#2 下面是我想分的段，因为我想分3段，所以要有4个数
q_list=[0,0.25,0.5,1]

#3 使用qcut来cut
pd.qcut(ser,q=q_list)

#4 观察下面的结果
# 第一部分是告诉我每个学生所处的分数段，所以list里面有几个元素，就有几个range
# 第二部分是告诉我有3个interval，分别是多少，他们是按照我的要求cut的
"""
[(58.999, 72.5], (58.999, 72.5], (72.5, 83.5], (83.5, 99.0], (83.5, 99.0],
    (72.5, 83.5], (83.5, 99.0], (83.5, 99.0]]
Categories (3, interval[float64]): [(58.999, 72.5] < (72.5, 83.5] < (83.5, 99.0]]
"""
#——————————————————Example 2: tell me your range label ———————————————————
# 上面的方法是告诉我你的range，但是我可能更关心是是在25%~75%，而不是关心(72.5,83.5]
# 所以这里我们给他们加上label

ser=[59,62,83,84,95,76,89,99]
q_list=[0,0.25,0.5,1]
label_list=['bottom 25%','middle 50%','top 25%'] # 重点是要加上label
pd.qcut(ser,q=q_list,labels=label_list)

#观察下面结果
"""
[bottom 25%, bottom 25%, middle 50%, top 25%, top 25%, middle 50%, top 25%, top 25%]
Categories (3, object): [bottom 25% < middle 50% < top 25%]"""

#—————————————————Example 3: Apply qcut to DataFrame —————————————————————
# 最有用的是：我有学生名字和他们的成绩，然后得到谁在什么range，这就要设计DF
#1 prepare a DF
score_list=[59,62,83,84,95,76,89,99]
student_list=['Amy','Bob','Cathy','Dodge','Edan','Fred','Grey','Helene']
DF=pd.DataFrame({'student':student_list,'score':score_list})
"""
  student  score
0     Amy     59
1     Bob     62
2   Cathy     83
3   Dodge     84
4    Edan     95
5    Fred     76
6    Grey     89
7  Helene     99"""

#2 apply qcut to a column and save results to a new column
q_list=[0,0.25,0.5,1]
label_list=['bottom 25%','middle 50%','top 25%'] # 重点是要加上label
DF['band']=pd.qcut(DF['score'],q=q_list,labels=label_list)
"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
L03 Set
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

"""********************************************************************
1.Create a set

********************************************************************"""
#1 create a set
mySet={1,2,3}

#2 create an empty set
mySet=set({})
mySet=set()

"""********************************************************************
2. Compute set differences or union or intersection
********************************************************************"""
A={1,2,3,4}
B={3,4,5,6}

A-B     #--> difference {1, 2}
A|B     #--> union {1, 2, 3, 4, 5, 6}
A&B     #--> intersection {3, 4}

"""********************************************************************
3. subset
syntax:
    SET.issubset(SET)
********************************************************************"""

A={1,2}
B={1,2,3,4}
C=set()

A.issubset(B) #Out[]: True
B.issubset(A) #Out[]: False
C.issubset(B) #Out[]: True

#空集是别人的subset.这是个需要注意的地方，如果我的set1是空的，结果也是True，
#但是事实就是set1并没有用set2中的元素，不应该和其他True的情况混为一谈。
#我可以想到的解决办法就是用len(set1)==0来把空集的情况摘出去。

"""********************************************************************
4. SET's statistics
********************************************************************"""
mySet={1,2,3,4}
sum(mySet) #--> 10
min(mySet) #--> 1

import statistics as stat
stat.stdev(mySet) #--> 1.290994

"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
L04 Numpy 101
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

import numpy as np

"""***********************************************************************
1. Create a numpy array
***********************************************************************"""
#————————————1.create an array by manual input one by one—————————————————
np.array([1,3,4,5])
"Out[]: array([1, 3, 4, 5])"

np.array(['Amy','Bob','Cathy'])
"Out[]: array(['Amy', 'Bob', 'Cathy'], dtype='<U5')"

#————————————2.Use np.arange() to create number array—————————————————————
"""Syntax:   
    arange([start,] stop[, step,], dtype=None)
    Return evenly spaced values within a given interval."""

np.arange(3)    #    array([0, 1, 2])    
np.arange(3.0)  #   array([ 0.,  1.,  2.])
np.arange(3,7)  #   array([3, 4, 5, 6])
np.arange(3,7,2)#   array([3, 5])

#———————————3.Use np.random.randint() to create random integer————————————
np.random.randint(10,100,5)

"""***********************************************************************
enumerate的用法
***********************************************************************"""
#enumerate()代表的是一组list的index和value组成的tuple

myList=[2,3,4,5]

for idx,val in enumerate(myList):
    print (idx,val)
"""
0 2
1 3
2 4
3 5
"""

for item in enumerate(myList):
    print (item)

"""
(0, 2)
(1, 3)
(2, 4)
(3, 5)"""

myDict={'a':1,'b':2,'c':3}

# dictionary也可以用，但是返回的是顺序和index没有value

for idx,val in enumerate(myDict):
    print (idx,val)
"""
0 a
1 b
2 c"""

for item in enumerate(myDict):
    print (item)
"""
(0, 'a')
(1, 'b')
(2, 'c')"""

"""********************************************************************
Question 1
Level 1

Question:
Write a program which will find all such numbers which are divisible 
by 7 but are not a multiple of 5, between 2000 and 3200 (both included).
The numbers obtained should be printed in a comma-separated sequence 
on a single line.

********************************************************************"""
output=[]
for i in range(2000,3201):
    if ((i%7==0) & (i%5!=0)):
        output.append(i)
print (output)


"""********************************************************************
Question 2
Level 1

Question:
Write a program which can compute the factorial （阶乘） of a given numbers.
The results should be printed in a comma-separated sequence on a single line.
Suppose the following input is supplied to the program:
8
Then, the output should be:
40320

********************************************************************"""
#1 这是我自己写的，我没有考虑0！的情况
def give_me_factorial(num):
    myFactor=1
    for i in range(1,num+1):
        myFactor=myFactor*i
    return myFactor

give_me_factorial(8)

#2 这是答案，这个是很有启发的-一个function引用自己

def give_me_factorial(num):
    if num==0:
        result=1
    else:
        result=num*give_me_factorial(num-1)
    return result

give_me_factorial(8)

"""********************************************************************
Question 3
Level 1

Question:
With a given integral number n, write a program to generate a dictionary
 that contains (i, i*i) such that is an integral number between 1 and n 
 (both included). and then the program should print the dictionary.
Suppose the following input is supplied to the program:
8
Then, the output should be:
{1: 1, 2: 4, 3: 9, 4: 16, 5: 25, 6: 36, 7: 49, 8: 64}

********************************************************************"""
def give_me_square(num):
    output={}
    for i in range(1,num+1):
        output[i]=i**2
    print (output)

give_me_square(8)

"""********************************************************************
Question 4
Level 1

Question:
Write a program which accepts a sequence of comma-separated numbers from
 console and generate a list and a tuple which contains every number.
Suppose the following input is supplied to the program:
34,67,55,33,12,98
Then, the output should be:
['34', '67', '55', '33', '12', '98']
('34', '67', '55', '33', '12', '98')

********************************************************************"""
def give_me_list_tuple(myStr):
    myList=myStr.split(',')
    myTuple=tuple(myList)
    print(myList)
    print(myTuple)

give_me_list_tuple('34,67,55,33,12,98')

"""********************************************************************
Question 5
Level 1

Question:
Define a class which has at least two methods:
getString: to get a string from console input
printString: to print the string in upper case.
Also please include simple test function to test the class methods.

********************************************************************"""
# 这题我不会做，class的没有学好

class InputOutString(object):
    def __init__(self):
        self.s = ""

    def getString(self):
        self.s = input()

    def printString(self):
        print (self.s.upper())

strObj = InputOutString()
strObj.getString()
strObj.printString()

"""********************************************************************
Question 6
Level 2

Question:
Write a program that calculates and prints the value according to the 
given formula:
    Q = Square root of [(2 * C * D)/H]

Following are the fixed values of C and H:
    C is 50. 
    H is 30.
    D is the variable whose values should be input to your program in a 
        comma-separated sequence.

Example:
Let us assume the following comma separated input sequence is given to 
the program:
    100,150,180

The output of the program should be:
    18,22,24
********************************************************************"""
#这里有2个点： list的one-line loop-through； int,str之间的转换
def Calc_Q():
    def Q(D):                          #1 Define the formula
        C=50
        H=30
        D=int(D)
        return str(int(((2 * C * D)/H)**(0.5)))
    D_list=(input()).split(",")        #2 Require input
    output_list=[Q(D) for D in D_list] #3 loop through input to compute
    return (",").join(output_list)     #4 join the output list

Calc_Q()

"""********************************************************************
Question 7
Level 2

Question:
Write a program which takes 2 digits, X,Y as input and generates a 
2-dimensional array. The element value in the i-th row and j-th column 
of the array should be i*j.

Note: i=0,1.., X-1; j=0,1,¡­Y-1.

Example
Suppose the following inputs are given to the program:
3,5
Then, the output of the program should be:
[[0, 0, 0, 0, 0], [0, 1, 2, 3, 4], [0, 2, 4, 6, 8]] 
********************************************************************"""
#这题的重点是nested list
def create_nest():
    myinput=input()
    input_list=myinput.split(",")
    X=int(input_list[0])
    Y=int(input_list[1])
    return [[x*y for y in range(Y)] for x in range(X)]

create_nest()
"""********************************************************************
Question 8
Level 2

Question:
Write a program that accepts a comma separated sequence of words as input
 and prints the words in a comma-separated sequence after sorting them 
 alphabetically.

Suppose the following input is supplied to the program:
without,hello,bag,world

Then, the output should be:
bag,hello,without,world
********************************************************************"""
def sort_words():
    myinput=input()
    input_list=myinput.split(",")
    input_list.sort()
    return ",".join(input_list)

sort_words()

"""********************************************************************
Question 9
Level 2

Question:
Write a program that accepts sequence of lines as input and prints the 
lines after making all characters in the sentence capitalized.

Suppose the following input is supplied to the program:
Hello world
Practice makes perfect

Then, the output should be:
HELLO WORLD
PRACTICE MAKES PERFECT
********************************************************************"""
#这题我开始不会做，因为不知道怎么接收分行的Input。一回车就结束了啊。所以学习一下
lines = []
while True:
    s = input()
    if s:
        lines.append(s.upper())
    else:
        break;

for sentence in lines:
    print (sentence)

"""********************************************************************
Question 10
Level 2

Question:
Write a program that accepts a sequence of whitespace separated words as
 input and prints the words after removing all duplicate words and sorting
 them alphanumerically.

Suppose the following input is supplied to the program:
hello world and practice makes perfect and hello world again

Then, the output should be:
again and hello makes perfect practice world
********************************************************************"""
myinput="maomao is a good cat"
myinput_set=set(myinput.split(" "))
myinput_list=list(myinput_set)
myinput_list.sort()

for word in myinput_list:
    print(word,end=" ")

"""********************************************************************
Question 11
Level 2

Question:
Write a program which accepts a sequence of comma separated 4 digit binary
 numbers as its input and then check whether they are divisible by 5 or 
 not. The numbers that are divisible by 5 are to be printed in a comma 
 separated sequence.

Example:

    0100,0011,1010,1001

Then the output should be:

    1010
Notes: Assume the data is input by console.
********************************************************************"""
def give_me_five(myStr):

    myList=myStr.split(",")
    my5List=[x for x in myList if (x[0]!='0')&((int(x))%5==0)]
    return ",".join(my5List)

give_me_five('0100,0011,1010,1001')

"""********************************************************************
Question 12
Level 2

Question:
Write a program, which will find all such numbers between 1000 and 3000
 (both included) such that each digit of the number is an even number.
The numbers obtained should be printed in a comma-separated sequence on
 a single line. - QW: I don't want it in a line, I want a list
********************************************************************"""
 
def give_me_even(start,end):
    def all_even(num):
        return all(int(x)%2==0 for x in str(num)) 
    output=[x for x in range(start,end+1) if all_even(x)]
    return output    
give_me_even(1000,3000)

"""********************************************************************
Question 13
Level 2

Question:
Write a program that accepts a sentence and calculate the number of 
letters and digits.

Suppose the following input is supplied to the program:
hello world! 123

Then, the output should be:
LETTERS 10
DIGITS 3

********************************************************************"""
import re
def give_me_count(myStr):
    letter_result=re.compile(r'[a-zA-Z]').findall(myStr)
    digit_result=re.compile(r'[0-9]').findall(myStr)
    print('LETTERS {}'.format(len(letter_result)))
    print('DIGITS {}'.format(len(digit_result)))

give_me_count('hello world! 123 mao 3 4 ')

"""********************************************************************
Question 14
Level 2

Question:
Write a program that accepts a sentence and calculate the number of upper
 case letters and lower case letters.
Suppose the following input is supplied to the program:
Hello world!
Then, the output should be:
UPPER CASE 1
LOWER CASE 9

********************************************************************"""
def give_me_letter(myStr):
    letter_result=re.compile(r'[A-Z]').findall(myStr)
    digit_result=re.compile(r'[a-z]').findall(myStr)
    print('UPPER CASE {}'.format(len(letter_result)))
    print('LOWER CASE {}'.format(len(digit_result)))

give_me_letter('Hello world!')

"""********************************************************************
Question 15
Level 2

Question:
Write a program that computes the value of a+aa+aaa+aaaa with a given 
digit as the value of a.

Suppose the following input is supplied to the program:
9

Then, the output should be:
11106

********************************************************************"""
a=7
total=(1+11+111+1111)*a
total

"""********************************************************************
Question 16
Level 2

Question:
Use a list comprehension to square each odd number in a list. 
The list is input by a sequence of comma-separated numbers.

Suppose the following input is supplied to the program:
1,2,3,4,5,6,7,8,9

Then, the output should be:
1,3,5,7,9
********************************************************************"""
def give_me_odd(myInput):
    myinputList=myInput.split(",")
    myoutputList=[x for x in myinputList if int(x)%2==1]
    return ",".join(myoutputList)
give_me_odd('1,21,33,4,5,62,7,8,9')

"""********************************************************************
Question 17
Level 2

Question:
Write a program that computes the net amount of a bank account based a 
transaction log from console input. The transaction log format is shown 
as following:

D 100
W 200
D means deposit while W means withdrawal.

Suppose the following input is supplied to the program:
D 300
D 300
W 200
D 100

Then, the output should be:
500
********************************************************************"""
def compute_net_bal():
    myInput=[]
    while True:
        s=input()
        if s:
            myInput.append(s)
        else:
            break
    
    def give_me_amt(entry):
        sign,amt=entry.split(" ")
        if sign=='D':
            return int(amt)
        else:
            return (-1)*int(amt)
    
    myAmt=[give_me_amt(x) for x in myInput]
    return sum(myAmt)

compute_net_bal()
"""********************************************************************
Question 18
Level 3

Question:
A website requires the users to input username and password to register. 
Write a program to check the validity of password input by users.

Following are the criteria for checking the password:
1. At least 1 letter between [a-z]
2. At least 1 number between [0-9]
1. At least 1 letter between [A-Z]
3. At least 1 character from [$#@]
4. Minimum length of transaction password: 6
5. Maximum length of transaction password: 12

Your program should accept a sequence of comma separated passwords and 
will check them according to the above criteria. Passwords that match 
the criteria are to be printed, each separated by a comma.

Example
If the following passwords are given as input to the program:
ABd1234@1,a F1#,2w3E*,2We3345

Then, the output of the program should be:
ABd1234@1
********************************************************************"""

def good_pw(pw):
    cond1=len(re.compile('[a-z]').findall(pw))
    cond2=len(re.compile('[A-Z]').findall(pw))    
    cond3=len(re.compile('[0-9]').findall(pw))    
    cond4=len(re.compile('[$#@!]').findall(pw)) 
    cond5=len(pw)>5
    cond6=len(pw)<13
    
    if cond1*cond2*cond3*cond4*cond5*cond6>0:
        return True
    else:
        return False

def pick_good_pw():
    myInput=input()
    myInput_list=myInput.split(",")
    myOutput_list=[x for x in myInput_list if good_pw(x)]
    print (",".join(myOutput_list))
    
pick_good_pw()
"""********************************************************************
Question 19
Level 3

Question:
You are required to write a program to sort the (name, age, height) tuples
 by ascending order where name is string, age and height are numbers. 
 The tuples are input by console. The sort criteria is:
1: Sort based on name;
2: Then sort based on age;
3: Then sort by score.

The priority is that name > age > score.
If the following tuples are given as input to the program:

Tom,19,80
John,20,90
Jony,17,91
Jony,17,93
Json,21,85

Then, the output of the program should be:
[('John', '20', '90'), ('Jony', '17', '91'), ('Jony', '17', '93'), 
('Json', '21', '85'), ('Tom', '19', '80')]
********************************************************************"""
#1 下面是我的做法，用df的sort的方法，不是很好，但是可以看看
import pandas as pd
myInput=[]

while True:
    s=input()
    if s:
        myInput.append(s)
    else:
        break

data=[x.split(",") for x in myInput]
df=pd.DataFrame(data,columns=['name','age','score'])
df=df.sort_values(['name','age','score'])
nest_list=df.values.tolist()
output=[tuple(i) for i in nest_list]

#2 这是一个multiple sorting的方法，需要学习
from operator import itemgetter
data.sort(key=itemgetter(0,1,2))
"""[['john', '20', '90'], ['jony', '17', '91'], ['tom', '18', '80']]"""

data.sort(key=itemgetter(2,1,0))
"""[['tom', '18', '80'], ['john', '20', '90'], ['jony', '17', '91']]"""

"""********************************************************************
Question 20
Level 3

Question:
Define a class with a generator which can iterate the numbers, 
which are divisible by 7, between a given range 0 and n.

********************************************************************"""
class Q20(object):
    def __init__(self):
        pass
    def give_me_seven(self,n):
        self.n=n
        myList=[x for x in range(self.n+1) if x%7==0]
        print (myList, sep=",") # bookmark pythonic printing
        print (myList)
        print (*myList)
        print (*myList,sep='\t')
    
mytest=Q20()    
mytest.give_me_seven(100)

"""********************************************************************
Question 21
Level 3

Question:
A robot moves in a plane starting from the original point (0,0). 
The robot can move toward UP, DOWN, LEFT and RIGHT with a given steps. 
The trace of robot movement is shown as the following:

UP 5,DOWN 3,LEFT 3,RIGHT 2

The numbers after the direction are steps. 
Please write a program to compute the distance from current position 
after a sequence of movement and original point. 
If the distance is a float, then just print the nearest integer.

Example:
If the following tuples are given as input to the program:
UP 5
DOWN 3
LEFT 3
RIGHT 2

Then, the output of the program should be:
2
********************************************************************"""
class movement(object):
    def __init__(self):
        self.myMove={'UP':0,'DOWN':0,'LEFT':0,'RIGHT':0}
    
    def get_movement(self):
        myMove=self.myMove
        while True:
            s=input()
            if s:
                mySplit=s.split(" ")
                myDir=mySplit[0]
                myLength=int(mySplit[1])
                myMove[myDir]=myMove[myDir]+myLength
            else:
                break
        self.myMove=myMove    
        
    def get_location(self):
        myLoc={'X':0,'Y':0}
        myLoc['Y']=myLoc['Y']+self.myMove['UP']-self.myMove['DOWN']
        myLoc['X']=myLoc['X']+self.myMove['RIGHT']-self.myMove['LEFT']
        self.loc=myLoc
        
    def get_distance(self):
        myLoc=self.loc
        raw_dist=((myLoc['X'])**2+(myLoc['Y'])**2)**(0.5)
        self.distance=int(round(raw_dist,0))
        
### Testing    
mytest=movement()
mytest.get_movement() # run method to provide input
mytest.myMove #{'UP': 10, 'DOWN': 2, 'LEFT': 7, 'RIGHT': 1}
mytest.get_location()
mytest.loc #{'X': -6, 'Y': 8}
mytest.get_distance()
mytest.distance #-->10

"""********************************************************************
Question 22
Level 3

Question:
Write a program to compute the frequency of the words from the input. 
The output should output after sorting the key alphanumerically. 
Suppose the following input is supplied to the program:

    New to Python or choosing between Python 2 and Python 3? 
    Read Python 2 or Python 3.

Then, the output should be:
2:2
3.:1
3?:1
New:1
Python:5
Read:1
and:1
between:1
choosing:1
or:2
to:1
********************************************************************"""
def count_words(myStr):
    myDict={}
    for each in myStr.split(" "):
        myDict.setdefault(each,0)
        myDict[each]+=1
    
    for each in sorted(myDict.items()):
        print (str(each[0])+":"+str(each[1]))

myStr="""New to Python or choosing between Python 2 and Python 3? Read Python 2 or Python 3."""
count_words(myStr)

"""********************************************************************
Question 23
level 1

Question:
    Write a method which can calculate square value of number
********************************************************************""" 
def comp_square(num):
    return num**2

comp_square(5)

"""********************************************************************
Question 24
Level 1

Question:
    Python has many built-in functions, and if you do not know how to use
    it, you can read document online or find some books. But Python has a
    built-in document function for every built-in functions.
    
    Please write a program to print some Python built-in functions 
    documents, such as abs(), int(), raw_input()
    And add document for your own function
********************************************************************""" 
print (abs.__doc__)
print (int.__doc__)
print (input.__doc__)
   
def square(num):
    '''Return the square value of the input number.
    
    The input number must be integer.
    '''
    return num ** 2

print (square.__doc__)


def maomi_test(catname):
    """ All the cats are good cats"""
    print (catname + " is a good cat.")

maomi_test('maomao')
print(maomi_test.__doc__)

"""********************************************************************
Question 25 bookmark - new knowledge
Level 1

Question:
    Define a class, which have a class parameter and have a same instance
    parameter.
********************************************************************"""
class Person:
    # Define the class parameter "name"
    name = "Person"
    
    def __init__(self, name = None):
        # self.name is the instance parameter
        self.name = name

#1 Testing 1
jeffrey = Person("Jeffrey")
Person.name
jeffrey.name
print ("{} name is {}".format(Person.name, jeffrey.name))

#2 Testing 2
nico = Person()
nico.name = "Nico"
print ("{} name is {}".format(Person.name, nico.name))

"""********************************************************************
Question26:
Define a function which can compute the sum of two numbers.

********************************************************************"""

def mySum(a,b):
    return a+b

mySum(100,10)

"""********************************************************************
Question27:
Define a function that can convert a integer into a string and print it in console.

********************************************************************"""
def make_it_str(myInt):
    print (str(myInt))

make_it_str(56)

"""********************************************************************
Question28:
Question:
Define a function that can receive two integral numbers in string form 
and compute their sum and then print it in console.

********************************************************************"""

def mySum():
    myInput=input().split(",")
    mySum=int(myInput[0])+int(myInput[1])
    print (mySum)

"""********************************************************************
Question 29:
Define a function that can accept two strings as input and concatenate
 them and then print it in console.
********************************************************************"""
def myconcat(a,b):
    print (str(a)+str(b))

myconcat('mao','mi')

"""********************************************************************
Question 30:
Define a function that can accept two strings as input and print the 
string with maximum length in console. If two strings have the same 
length, then the function should print al l strings line by line.
********************************************************************"""
#def print_max(a,b):
a='maom'
b='duduo'

if len(a)>len(b):
    print (a)
elif len(a)==len(b):
    print (a + '\n' +b)
else:
    print (b)

"""********************************************************************
Question 31
Define a function that can accept an integer number as input and print
 the "It is an even number" if the number is even, otherwise print 
 "It is an odd number".

********************************************************************"""
num=9

if num%2==0:
    print ('It is an even number.')
else:
    print ('It is an odd number.')

"""********************************************************************
Question 32,33,34,35
Define a function which can print a dictionary where the keys are numbers
 between 1 and 3 (both included) and the values are square of keys.
********************************************************************""" 

mytest={1:34,2:4,3:25,4:16}   
output={}
for each in mytest.keys():
    if (each in range(1,10)) & (mytest[each]==each**2):
        output.setdefault(each,mytest[each])
output #export the entire dict
output.values() #export values only
output.keys()   # export keys only

"""********************************************************************
Question 36,37,38,39,40,41

Question:
Define a function which can generate and print a list where the values 
are square of numbers between 1 and 20 (both included).

********************************************************************""" 
myList=[x**2 for x in range(1,21)] # generate a list
sorted(myList)[-5::] #export the largest 5
sorted(myList)[0:5] #export the smallest 5
myList[-5::] # export the last 5
myList[5::] # export all but the first 5
tuple(myList) # export in tuple

"""********************************************************************
Question 42

Question:
With a given tuple (1,2,3,4,5,6,7,8,9,10), write a program to print the 
first half values in one line and the last half values in one line. 

********************************************************************""" 
myTuple=(1,2,3,4,5,6,7,8,9,10)
halfLen=int(round(len(myTuple)/2,0))
",".join([str(x) for x in myTuple[0:halfLen]])
",".join([str(x) for x in myTuple[halfLen::]]) 
#bookmark - join values not string - use one-line list

"""********************************************************************
Question 43

Question:
Write a program to generate and print another tuple whose values are even
 numbers in the given tuple (1,2,3,4,5,6,7,8,9,10). 
********************************************************************""" 
myTuple=(1,2,3,4,5,6,7,8,9,10)
tuple([x for x in myTuple if x%2==0])


"""********************************************************************
Question 44
Write a program which accepts a string as input to print "Yes" if the 
string is "yes" or "YES" or "Yes", otherwise print "No". 

********************************************************************""" 
myInput='yEs'
if myInput in ['yes','YES','Yes']:
    print ('Yes')
else:
    print ('No')
    
"""********************************************************************
Question 45,46,47,48,49
Write a program which can filter even numbers in a list by using filter
 function. The list is: [1,2,3,4,5,6,7,8,9,10].

********************************************************************""" 
myList=[1,2,3,4,5,6,7,8,9,10]
[x for x in myList if x%2==0] # filter for even numbers
[x**2 for x in myList] # get squared
[x**2 for x in myList if x%2==0] # get square for even numbers


"""********************************************************************
Question 50
Define a class named American which has a static method called 
printNationality.
Use @staticmethod decorator to define class static method. bookmark
********************************************************************""" 
class American(object):
    @staticmethod
    def printNationality():
        print ("American")

anAmerican = American()
anAmerican.printNationality()
American.printNationality()
"""********************************************************************
Question 51

Define a class named American and its subclass NewYorker. bookmark

********************************************************************""" 
class American(object):
    pass
class NewYorker(American):
    pass
anAmerican = American()
aNewYorker = NewYorker()
print (anAmerican)
print (aNewYorker)

"""********************************************************************
Question 52

Define a class named Circle which can be constructed by a radius. 
The Circle class has a method which can compute the area. 

********************************************************************"""
#1 IF I set the radius = 2
class Circle(object):
    def __init__(self,r):
        self.radius=r
    def area(self):
        return (self.radius)**2*3.1415926

myCircle=Circle(4)
myCircle.radius
myCircle.area()

"""********************************************************************
Question 53
Define a class named Rectangle which can be constructed by a length 
and width. The Rectangle class has a method which can compute the area. 
********************************************************************"""
class Rectangle(object):
    def __init__(self,a,b):
        self.a=a
        self.b=b
    def area(self):
        return self.a*self.b
    
myR=Rectangle(3,4)
myR.area() #--> 12

"""********************************************************************
Question 54
 Define a class named Shape and its subclass Square. The Square class 
 has an init function which takes a length as argument. Both classes
 have a area function which can print the area of the shape where Shape's
 area is 0 by default.

********************************************************************"""

class Shape(object):
    def __init__(self):
        pass
    def area(self):
        return 0
class Square(Shape):
    def __init__(self,l):
        self.l=l
    def area(self):
        return (self.l)**2

myS=Shape()
myS.area()
yourS=Square(2)
yourS.l
yourS.area()

"""********************************************************************
Question 54
Please raise a RuntimeError exception. bookmark
********************************************************************"""

raise RuntimeError('Maomao is the best')

"""********************************************************************
Question 55
Write a function to compute 5/0 and use try/except to catch the exceptions.
********************************************************************"""
try:
    5/0
except ZeroDivisionError:
    print ("cannot be divided by zero")
"""********************************************************************
Question 56
Define a custom exception class which takes a string message as attribute.
这个我不懂 bookmark
********************************************************************"""
class MyError(Exception):
    """My own exception class

    Attributes:
        msg  -- explanation of the error
    """

    def __init__(self, msg):
        self.msg = msg

error = MyError("something wrong")

"""********************************************************************
Question 57,58
Assuming that we have some email addresses in the 
"username@companyname.com" format, please write program to print the 
user name of a given email address. Both user names and company names
 are composed of letters only.

Example:
If the following email address is given as input to the program:

john@google.com

Then, the output of the program should be:

john

In case of input data being supplied to the question, it should be
 assumed to be a console input.
********************************************************************"""
myEmail='john@google.com'
(myEmail.split("@"))[0]
((myEmail.split("@"))[1]).split(".")[0]

#如果我用regular expression实现
import re
result=re.compile('(\w+)@(\w+).com').search(myEmail)
result.group(1) #--> 'john'
result.group(2) #--> 'google'

"""********************************************************************
Question 59
Write a program which accepts a sequence of words separated by whitespace
 as input to print the words composed of digits only.

Example:
If the following words is given as input to the program:

2 cats and 3 dogs.

Then, the output of the program should be:

['2', '3']

In case of input data being supplied to the question, it should be 
assumed to be a console input.

********************************************************************"""
myStr='2 cats and 3 dogs.'
#1 这是我loop through的方法，不是很聪明
myList=[]
for x in myStr:
    try:
        myList.append(str(int(x)))
    except:
        continue
myList

#2 我没有想到这里可以用regular expression
re.compile('[0-9]+').findall(myStr)
re.compile('\d+').findall(myStr) 

"""********************************************************************
Question 60
Print a unicode string "hello world". bookmark(unicode怎么和其他不一样)

********************************************************************"""
myStr=u'2 cats and 3 dogs.'
print (myStr)

"""********************************************************************
Question 61

Write a program to read an ASCII string and to convert it to a unicode
 string encoded by utf-8. bookmark

********************************************************************"""
s = input()
u = unicode(s ,"utf-8") # I think this has been decommissioned. to check
print (u)

"""********************************************************************
Question 62
Write a special comment to indicate a Python source code file is in 
unicode.

********************************************************************"""
# -*- coding: utf-8 -*-

"""********************************************************************
Question 63
Write a program to compute 1/2+2/3+3/4+...+n/n+1 with a given n input 
by console (n>0).

Example:
If the following n is given as input to the program:
5

Then, the output of the program should be:
3.55
********************************************************************"""
n=5
sum([x/(x+1) for x in range(1,n+1)])

"""********************************************************************
Question 64

Write a program to compute:
    f(n)=f(n-1)+100 when n>0
    and f(0)=1

with a given n input by console (n>0).
Example:
If the following n is given as input to the program:
5

Then, the output of the program should be:
500
********************************************************************"""
def calc(n):
    if n>0:
        result=calc(n-1)+100
    if n==0:
        result=0
    return result
    
calc(5)

#这是我上面学的函数自体循环

"""********************************************************************
Question 65
The Fibonacci Sequence is computed based on the following formula:


f(n)=0 if n=0
f(n)=1 if n=1
f(n)=f(n-1)+f(n-2) if n>1

Please write a program to compute the value of f(n) with a given.

Example:
If the following n is given as input to the program:
7

Then, the output of the program should be:
13
********************************************************************"""
def Fi(n):
    if n==0:result=0
    elif n==1:result=1
    else:result=Fi(n-1)+Fi(n-2)
    return result

Fi(7)
# beautiful style
"""********************************************************************
Question 66
The Fibonacci Sequence is computed based on the following formula:

f(n)=0 if n=0
f(n)=1 if n=1
f(n)=f(n-1)+f(n-2) if n>1

Please write a program using list comprehension to print the Fibonacci
 Sequence in comma separated form with a given n input by console.

Example:
If the following n is given as input to the program:
7
Then, the output of the program should be:
0,1,1,2,3,5,8,13
********************************************************************"""
def printFi(n):
    def Fi(n):
        if n==0:result=0
        elif n==1:result=1
        else:result=Fi(n-1)+Fi(n-2)
        return result    
    return ",".join([str(Fi(i)) for i in range(n+1)])
        
printFi(7)

"""********************************************************************
Question 67

Please write a program using generator to print the even numbers between
 0 and n in comma separated form while n is input by console.

Example:
If the following n is given as input to the program:
10

Then, the output of the program should be:
0,2,4,6,8,10

********************************************************************"""
n=10
",".join([str(i) for i in range(n+1) if i%2==0])


"""********************************************************************
Question 68

Please write a program using generator to print the numbers which can
 be divisible by 5 and 7 between 0 and n in comma separated form while
 n is input by console.

Example:
If the following n is given as input to the program:

100
Then, the output of the program should be:

0,35,70
********************************************************************"""
n=200
",".join([str(i) for i in range(n+1) if ((i%5==0)&(i%7==0))])

"""********************************************************************
Question 69

Please write assert statements to verify that every number in the list
 [2,4,6,8] is even.

********************************************************************"""
all(i%2==0 for i in [2,4,6,8])

"""********************************************************************
Question 70
Please write a program which accepts basic mathematic expression from
 console and print the evaluation result.

Example:
If the following string is given as input to the program:
35+3

Then, the output of the program should be:
38
********************************************************************"""
# bookmark 0 eval()
myInput='35+3'
eval(myInput)

"""********************************************************************
Question 71,72
Please write a binary search function which searches an item in a sorted
 list. The function should return the index of element to be searched in
 the list.

********************************************************************"""
from bisect import bisect
myList=[2,4,6,8,10]
bisect(myList,8) #-->4
bisect(myList,7) #-->3

myList=[2,4,6,8,10]
def tell_me_loc(n):
    for i in range(len(myList)):
        if myList[i]==n:
            return i
            break
    return 'Not found in the list.'
    
tell_me_loc(9)    


"""********************************************************************
Question 73,74
Please generate a random float where the value is between 10 and 100 
using Python math module.
********************************************************************"""
# bookmark random module
import random
random.random()*100 #random float where the value is between 10 and 100
random.random()*(95-5)+5 #random between 5 and 95


"""********************************************************************
Question 74
Please write a program to output a random even number between 0 and 10
 inclusive using random module and list comprehension.

********************************************************************"""
#下面是我写的
evenList=[x for x in range(0,11) if x%2==0]
evenList[int(round(random.random()*len(evenList),0))]

#random.choice()才是正解 bookmark
evenList=[x for x in range(0,11) if x%2==0]
random.choice(evenList)

"""********************************************************************
Question 75
Please write a program to output a random number, which is divisible
 by 5 and 7, between 0 and 100 inclusive using random module and list 
 comprehension.

********************************************************************"""
myList=[x for x in range(0,101) if ((x%5==0)&(x%7==0))]
random.choice(myList)

"""********************************************************************
Question 76
Please write a program to generate a list with 5 random numbers between
 100 and 200 inclusive.

********************************************************************"""
random.sample(range(100,201),5)


"""********************************************************************
Question 77
Please write a program to randomly generate a list with 5 even numbers
 between 100 and 200 inclusive.
********************************************************************"""
random.sample([x for x in range(100,201) if x%2==0],5)


"""********************************************************************
Question 78
Please write a program to randomly generate a list with 5 numbers, 
which are divisible by 5 and 7 , between 1 and 1000 inclusive.
********************************************************************"""
random.sample([x for x in range(1,1001) if ((x%5==0)&(x%7==0))],5)


"""********************************************************************
Question 79
Please write a program to randomly print a integer number between 7 
and 15 inclusive.

********************************************************************"""
random.random()*(15-7)+7 #我写的这个是错的，因为不是整数
# 
random.randrange(7,16)
#
random.choice(range(7,16))
"""********************************************************************
Question 80
Please write a program to compress and decompress the string 
"hello world!hello world!hello world!hello world!".

bookmark: zlib data compression我现在不是很明白，他给的答案也不靠谱
********************************************************************"""
import zlib


"""********************************************************************
Question 81
Please write a program to print the running time of execution of 
"1+1" for 100 times.

********************************************************************"""
#1这个是我写的
import time
t0=time.time()
i=0
while i<100:
    1+1
    i+=1
t1=time.time()
print (str(t1-t0)+" seconds")

#2这个是答案
from timeit import Timer #bookmark
t = Timer("for i in range(100):1+1")
print (t.timeit())


"""********************************************************************
Question 82
Please write a program to shuffle and print the list [3,6,7,8].

********************************************************************"""
from random import shuffle

myList=[3,6,7,8]
shuffle(myList) #shuffle() is a method, not a list
myList

"""********************************************************************
Question 83
Please write a program to generate all sentences where 
subject is in ["I", "You"] and 
verb is in ["Play", "Love"] and 
the object is in ["Hockey","Football"].
********************************************************************"""
sub_list=["I", "You"]
verb_list=["Play", "Love"]
object_list=["Hockey","Football"]

[x+" "+y+" "+z for x in sub_list for y in verb_list for z in object_list]

"""********************************************************************
Question 84
Please write a program to print the list after removing delete even 
numbers in [5,6,77,45,22,12,24].
********************************************************************"""
myList=[5,6,77,45,12,24]
newList=[x for x in myList if x%2!=0]
print (newList)

"""********************************************************************
Question 85
By using list comprehension, please write a program to print the list 
after removing delete numbers which are divisible by 5 and 7 in 
[12,24,35,70,88,120,155].
********************************************************************"""
myList=[12,24,35,70,88,120,155]
newList=[x for x in myList if ((x%5!=0) | (x%7!=0))]
print (newList)

"""********************************************************************
Question 85
By using list comprehension, please write a program to print the list 
after removing the 0th, 2nd, 4th,6th numbers in [12,24,35,70,88,120,155].
********************************************************************"""
myList=[12,24,35,70,88,120,155]
enumerate(myList) #bookmark -->结果是一组tuple (i,x)
enumerate.__doc__
[each for each in enumerate(myList)]

[x for (i,x) in enumerate(myList) if i not in [0,2,4,6]]

"""********************************************************************
Question 86

By using list comprehension, please write a program generate a 3*5*8 
3D array whose each element is 0. bookmark 3D array

********************************************************************"""
array = [[ [0 for col in range(8)] for col in range(5)] for row in range(3)]
print (array)

"""********************************************************************
Question 87
By using list comprehension, please write a program to print the list 
after removing the 0th,4th,5th numbers in [12,24,35,70,88,120,155].

********************************************************************"""
[x for (i,x) in enumerate([12,24,35,70,88,120,155]) if i not in [0,4,5]]


"""********************************************************************
Question 88
By using list comprehension, please write a program to print the list 
after removing the value 24 in [12,24,35,24,88,120,155].
********************************************************************"""
myList=[12,24,35,24,88,120,155]
myList.remove(24)
myList

"""********************************************************************
Question 89
With two given lists [1,3,6,78,35,55] and [12,24,35,24,88,120,155], 
write a program to make a list whose elements are intersection of the 
above given lists.
********************************************************************"""
a=[1,3,6,78,35,55]
b=[12,24,35,24,88,120,155]
list(set(a)&set(b))


"""********************************************************************
Question 90
With a given list [12,24,35,24,88,120,155,88,120,155], write a program
 to print this list after removing all duplicate values with original 
 order reserved.
********************************************************************"""
myList=[12,24,35,24,88,120,155,88,120,155]
result=[]
for x in myList:
    if x not in result:
        result.append(x)
result

"""********************************************************************
Question 91
Define a class Person and its two child classes: Male and Female. 
All classes have a method "getGender" which can print "Male" for Male 
class and "Female" for Female class.
********************************************************************"""
class Person(object):
    def getGender(self):
        print ('Unknown')
class Male(Person):
    def getGender(self):
        print ('Male')
class Female(Person):
    def getGender(self):
        print ('Female')

maomao=Female()
maomao.getGender()

"""********************************************************************
Question 92
Please write a program which count and print the numbers of each character
 in a string input by console.

Example:
If the following string is given as input to the program:

abcdefgabc

Then, the output of the program should be:

a,2
c,2
b,2
e,1
d,1
g,1
f,1

********************************************************************"""
myStr='abcdefgabc'
myDict={}
for each in myStr:
    myDict.setdefault(each,0)
    myDict[each]=myDict[each]+1

"""********************************************************************
Question 93
Please write a program which accepts a string from console and print it
 in reverse order.

********************************************************************"""
myStr='Please'
print (myStr[::-1])

"""********************************************************************
Question 94
Please write a program which accepts a string from console and print the
 characters that have even indexes.
********************************************************************"""
myStr='Please'
print (myStr[0::2])


"""********************************************************************
Question 95
Please write a program which prints all permutations of [1,2,3]
#bookmark
********************************************************************"""
import itertools
list(itertools.permutations([1,2,3]))


"""********************************************************************
Question 96
Write a program to solve a classic ancient Chinese puzzle: 
We count 35 heads and 94 legs among the chickens and rabbits in a farm. 
How many rabbits and how many chickens do we have?

********************************************************************"""
for i in range(35+1):
    for j in range(35+1):
        if (i+j==35) & (4*i+2*j==94):
            print ('There are {} rabbits and {} chickens.'.format(i,j))



"""********************************************************************
Question 
********************************************************************"""

"""********************************************************************
Question 
********************************************************************"""
"""********************************************************************
Question 
********************************************************************"""
"""***********************************************************************
1. How to import pandas and check the version?
[Assessment:I don't know]

***********************************************************************"""
import pandas as pd
pd.__version__ #--> '0.23.0' 这里考点是用__version__,__doc__的部分


"""***********************************************************************
2. How to create a series from a list, numpy array and dict? 
***********************************************************************"""

#1 create inputs first
myList=list('abcdefghijklmnopqrstuvwxyz')  #1 create a list
import numpy as np
myArray=np.arange(26)                      #2 create an array
myDict=dict(zip(myList,myArray))           #3 create a dictionary

#2 convert to a series
series1=pd.Series(myList)                          #1 from list
series2=pd.Series(np.arange(26))                   #2 from array
series3=pd.Series(myDict)                          #3 from dict

"""***********************************************************************
3. How to convert the index of a series into a column of a dataframe
***********************************************************************"""
#1 用上面的series3
series3
"""
a     0
b     1
c     2
d     3
e     4
f     5
g     6
h     7
i     8
j     9..."""

#2 我的想法是用pd.DataFrame
dataf1=pd.DataFrame(series3.index)

#3 答案是to_frame()
dataf2=series3.to_frame()


"""***********************************************************************
4. How to combine many series to form a dataframe? 
***********************************************************************"""
#1.input
series1=pd.Series([1,2,3])
series2=pd.Series([4,5,6])
series3=pd.Series([7,8,9])

#2. use concat combine, bookmark
pd.concat([series1,series2,series3],axis=1)
"""
   0  1  2
0  1  4  7
1  2  5  8
2  3  6  9"""

pd.concat([series1,series2,series3],axis=0)
"""
0    1
1    2
2    3
0    4
1    5
2    6
0    7
1    8
2    9"""

#3. use pd.DataFrame() bookmark
pd.DataFrame({'col1':series1,'col2':series2,'col3':series3})
"""
   col1  col2  col3
0     1     4     7
1     2     5     8
2     3     6     9"""

"""***********************************************************************
5. How to assign name to the series’ index? 
***********************************************************************"""
#1 下面是我写的，其实不对，我写的是如何添加index，题目的意思是给已有的index添加
#一个名字。比如我下面写的a,b,c是index，我要给它添加名字"letter"
series1=pd.Series([1,2,3])
series1.index=['a','b','c']
"""
a    1
b    2
c    3
dtype: int64"""

series2=pd.Series([1,2,3],index=['a','b','c'])
"""
a    1
b    2
c    3
dtype: int64"""

#2 下面是正解：注意"Name: Letter"
series1.name='Letter'
"""
a    1
b    2
c    3
Name: Letter, dtype: int64"""


"""***********************************************************************
6. How to get the items of series A not present in series B?
***********************************************************************"""
"""Input"""
ser1 = pd.Series([1, 2, 3, 4, 5]) 
ser2 = pd.Series([4, 5, 6, 7, 8])

#1. 我的想法是把他们存入set然后比较- 这个方法的问题是，我的结果没有保留series的
# 形式（如果不转换的话）
set1=set(ser1)
set2=set(ser2)
set1-set2
""" {1, 2, 3}"""

#2 这是给的答案- 一直于series的形式.利用isin，和DF一样使用boolean series
ser1[~ser1.isin(ser2)]


"""***********************************************************************
7. How to get the items not common to both series A and series B? 
我这里就不用set的方法了，复习isin的方法
***********************************************************************"""
ser1 = pd.Series([1, 2, 3, 4, 5]) 
ser2 = pd.Series([4, 5, 6, 7, 8])

#1下面是我用isin的方法
pd.concat([ser1[~ser1.isin(ser2)],ser2[~ser2.isin(ser1)]],axis=0)

#2答案其实很麻烦，但是提出了一个我不知道的np的function
import numpy as np
myIntersect=np.intersect1d(ser1,ser2)  # np intersect
myUnion=np.union1d(ser1,ser2)          # np union
myUnion[~myUnion.isin(myIntersect)]    # 这个会返回错误，因为myUnion是np array
#而不是series,正确的方法如下，要变成pd series，实在太麻烦
ser_u = pd.Series(np.union1d(ser1, ser2))  # union
ser_i = pd.Series(np.intersect1d(ser1, ser2))  # intersect
ser_u[~ser_u.isin(ser_i)]

"""***********************************************************************
8. How to get the minimum, 25th percentile, median, 75th, and max 
of a numeric series? 
***********************************************************************"""
#1 This is what I did based on my knowedge on DF
ser1 = pd.Series([1, 2, 3, 4, 5]) 
min(ser1)             #-->1
ser1.min()            #-->1
ser1.quantile(q=0.25) #-->2.0
ser1.median()         #-->3.0
ser1.max()            #-->5
ser1.quantile(q=0.75) #-->4.0

#2 The solution uses np.percentile
#这个思路是把max, min,median, 25%, 75%一起算出来，其实也不错
np.percentile(ser1,q=[0,25,50,75,100])

"""***********************************************************************
9. How to get frequency counts of unique items of a series? 
***********************************************************************"""
ser1 = pd.Series([1,2,3,3,4,5,5,5]) 
ser1.value_counts()
"""
5    3
3    2
4    1
2    1
1    1
dtype: int64"""

"""***********************************************************************
10. How to keep only top 2 most frequent values as it is and replace 
everything else as ‘Other’?
***********************************************************************"""
#1 This is my solution
ser1 = pd.Series([1,2,3,3,4,5,5,5]) 
otherlist=list((ser1.value_counts()).index)[2::]
for item in otherlist:
    ser1=ser1.replace(item,r'Other',regex=True)
ser1
"""
0    Other
1    Other
2        3
3        3
4    Other
5        5
6        5
7        5"""

#2 This is also my solution using one-line list
ser1 = pd.Series([1,2,3,3,4,5,5,5]) 
full_list=list(ser1)
otherlist=list((ser1.value_counts()).index)[2::]
["Other" if x in otherlist else x for x in full_list ]

"""***********************************************************************
11. How to bin a numeric series to 10 groups of equal size? 
***********************************************************************"""

import pandas as pd, numpy as np
ser = pd.Series(np.random.random(20)) 
print(ser) 
q_list=[0, .10, .20, .3, .4, .5, .6, .7, .8, .9, 1]
label_list=['1st','2nd','3rd','4th','5th','6th','7th','8th','9th','10th']
pd.qcut(ser, q=q_list,labels=label_list)
"""
0      2nd
1      4th
2      9th
3      6th
4     10th
5      8th
6      5th
7      3rd
8      7th
9      6th
10     5th
11    10th
12     2nd
13     7th
14     8th
15     1st
16     4th
17     3rd
18     9th
19     1st
dtype: category
Categories (10, object): [1st < 2nd < 3rd < 4th ... 7th < 8th < 9th < 10th]"""

# 知识点
#《1》 用random.random创造任意20个数字于0~1之间
#《2》用pd.qcut做分割
pd.qcut.__doc__
"Quantile-based discretization function"

ser=[1,2,3,4]
q_list=[0,0.25,0.5,1]
pd.qcut(ser, q=q_list)

ser=[1,2,3,4]
q_list=[0,0.25,0.5,1]
label_list=['1st','2nd','3rd']
pd.qcut(ser, q=q_list,labels=label_list)
 

"""***********************************************************************
12. How to convert a numpy array to a dataframe of given shape?
***********************************************************************"""
#1 create a np array first
import numpy as np
NPA=np.array([1,2,3,4]) #-array([1, 2, 3, 4])

#2 convert np array to a series
ser=pd.Series(NPA)

#3 concert series to a dataframe of given shape
pd.DataFrame(ser.values.reshape(2,2))
"""
   0  1
0  1  2
1  3  4"""

"""***********************************************************************
11. How to bin a numeric series to 10 groups of equal size? 
***********************************************************************"""

"""***********************************************************************
11. How to bin a numeric series to 10 groups of equal size? 
***********************************************************************"""

"""***********************************************************************
11. How to bin a numeric series to 10 groups of equal size? 
***********************************************************************"""

"""***********************************************************************
11. How to bin a numeric series to 10 groups of equal size? 
***********************************************************************"""

"""***********************************************************************
11. How to bin a numeric series to 10 groups of equal size? 
***********************************************************************"""

"""***********************************************************************
11. How to bin a numeric series to 10 groups of equal size? 
***********************************************************************"""

"""***********************************************************************
11. How to bin a numeric series to 10 groups of equal size? 
***********************************************************************"""




"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
X01_XlsxWriter 101
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

import os
import xlsxwriter
myPath=r'C:\Users\Danish\Desktop\python sample'
os.chdir(myPath)

"""********************************************************************
1. Create File_Tab_and Cell contents
syntax:
    xlsxwriter.Workbook()
    WB.add_worksheet()
    WS.write(location,value)
    WB.close()
********************************************************************"""
#1 create an Excel file
wb=xlsxwriter.Workbook('mytest.xlsx') # create workbook object
ws=wb.add_worksheet('mysheet')
ws.write('A1','Testing Summary')
wb.close()

"""********************************************************************
2. Create text and table
syntax:
    WS.write(location,value)
********************************************************************"""
wb=xlsxwriter.Workbook('mytest.xlsx') #1 create workbook object
ws=wb.add_worksheet('mysheet')        #2 create worksheet object
ws.write('A1','Testing Summary')      #3 write text on a sheet

expenses = ( ['Rent', 1000], 
            ['Gas', 100], 
            ['Food', 300], 
            ['Gym', 50])              #4 prepare a tuple for table

row=3;col=1                           #5 Set the table's location



for item,cost in expenses:
    ws.write(row,col,item)
    ws.write(row,col+1,cost)
    row+=1                            #6 loop-through to populate table

wb.close()                            #7 close the workbook

"""********************************************************************
. Create text and table
syntax:
    FORMAT=WB.add_format({'bold': True,'italic':True})
********************************************************************"""
wb=xlsxwriter.Workbook('mytest.xlsx') #1 create workbook object

bold=wb.add_format({'bold':True})     #2 create various format objects
bold_italic=wb.add_format({'bold':True,'italic':True})
money=wb.add_format({'num_format':'$#,##0'})
consolas=wb.add_format({'font_name':'consolas'})
                     
ws=wb.add_worksheet('mysheet')        #3 create worksheet object
ws.write('A1','Testing Summary',consolas) #4 write text on a sheet with format

expenses = ( ['Rent', 1000], 
            ['Gas', 100], 
            ['Food', 300], 
            ['Gym', 50])              #5 prepare a tuple for table

row=3;col=1                           #6 Set the table's location

for item,cost in expenses:
    ws.write(row,col,item)
    ws.write(row,col+1,cost,money)
    row+=1                            #6 loop-through to populate table

wb.close()                            #7 close the workbook
"""$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$
X01_XlsxWriter 102
$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$"""

import pandas as pd
import os
#import xlsxwriter
myPath=r'C:\Users\Danish\Desktop\python sample'
os.chdir(myPath)

"""********************************************************************
2. Export DF(s) to certain location
********************************************************************"""
df1=pd.DataFrame([[1,2],[3,4]])
df2=pd.DataFrame([[5,6],[7,8]])

writer=pd.ExcelWriter('temp.xlsx')
df1.to_excel(writer,sheet_name='mydf',startrow=2,startcol=2)
df2.to_excel(writer,sheet_name='mydf',startrow=2,startcol=5)
writer.close()

"""********************************************************************
1. Merge Cells
********************************************************************"""


"""********************************************************************
2. Insert a Chart on DF
********************************************************************"""


"""********************************************************************
3. Conditional Formatting
********************************************************************"""
